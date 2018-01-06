from openpyxl import load_workbook


wb = load_workbook(filename='empty_book.xlsx')
sheetnames = []

for sheet in wb.sheetnames:
    sheetnames.append(sheet)

sheet_ranges = wb['range names']

print('D2 cell value:', sheet_ranges['D2'].value)

print('Sheet: ', sheetnames[1], ' Value on F5', wb[sheetnames[1]]['F5'].value)

print('Sheet: ', sheetnames[2], ' Value on F5', wb[sheetnames[2]]['F5'].value)

print('Sheets: ', sheetnames)

